"""Tests for tabular-text + spreadsheet compression.

Covers detection (content_detector), the CSV→SmartCrusher bridge
(tabular_ingest), router wiring (content_router), and binary spreadsheet
ingestion (spreadsheet_ingest / compress_spreadsheet).
"""

from __future__ import annotations

import importlib.util

import pytest

from headroom.transforms.content_detector import (
    ContentType,
    DetectionResult,
    _is_md_separator,
    _looks_like_prose,
    _try_detect_delimited,
    _try_detect_markdown_table,
    detect_content_type,
)
from headroom.transforms.content_router import (
    CompressionStrategy,
    ContentRouter,
    ContentRouterConfig,
)
from headroom.transforms.tabular_ingest import (
    TabularCompressionResult,
    TabularCompressor,
    parse_csv,
    parse_fixed_width,
    parse_markdown_table,
    parse_tabular,
    to_records,
)

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


# Reusable fixtures ----------------------------------------------------------

CSV = "name,age,city\nAlice,30,NYC\nBob,25,LA\nCara,40,SF"
TSV = "id\tval\tnote\n1\ta\tx\n2\tb\ty\n3\tc\tz"
MARKDOWN = "| name | age |\n| --- | --- |\n| Alice | 30 |\n| Bob | 25 |\n| Cara | 40 |"


def _verbose_markdown(rows: int = 40) -> str:
    body = "\n".join(
        f"| user_{i} | {20 + i} | city_{i % 5} | active | engineering |" for i in range(rows)
    )
    return "| name | age | city | status | dept |\n| --- | --- | --- | --- | --- |\n" + body


# Detection ------------------------------------------------------------------


@pytest.mark.parametrize(
    "content,fmt",
    [(CSV, "csv"), (TSV, "csv"), (MARKDOWN, "markdown")],
)
def test_detects_tabular(content: str, fmt: str) -> None:
    result = detect_content_type(content)
    assert result.content_type is ContentType.TABULAR
    assert result.metadata.get("format") == fmt
    assert result.confidence >= 0.6


@pytest.mark.parametrize(
    "content,expected",
    [
        # Search output must not be stolen by tabular.
        (
            "src/main.py:42:def process():\nsrc/util.py:10:import os\nsrc/x.py:5:return 1",
            ContentType.SEARCH_RESULTS,
        ),
        # Build/log output stays a log.
        (
            "2026-01-01 INFO starting\n2026-01-01 WARN slow\n2026-01-01 ERROR boom",
            ContentType.BUILD_OUTPUT,
        ),
        # JSON arrays still go to the JSON path.
        ('[{"a": 1}, {"a": 2}, {"a": 3}]', ContentType.JSON_ARRAY),
        # Prose with incidental commas must NOT be tabular.
        (
            "Hello there, friend.\nThis is a sentence, yes.\nAnother line, ok.",
            ContentType.PLAIN_TEXT,
        ),
    ],
)
def test_does_not_misroute_to_tabular(content: str, expected: ContentType) -> None:
    assert detect_content_type(content).content_type is expected


# Detection — edge branches --------------------------------------------------


def test_is_md_separator_needs_two_columns() -> None:
    assert _is_md_separator("| --- | --- |")
    assert not _is_md_separator("| --- |")  # single column is not a separator
    assert not _is_md_separator("| a | b |")  # cells must be dashes


def test_markdown_table_needs_multiple_columns() -> None:
    # Valid separator below, but the header is a single column -> not a table.
    assert _try_detect_markdown_table(["x|", "---|---", "y|"]) is None


def test_delimited_needs_three_rows() -> None:
    assert _try_detect_delimited(["a,b,c", "1,2,3"]) is None


def test_delimited_rejects_delimiter_only_in_header() -> None:
    # Header has commas but the data rows don't: no stable column count.
    assert _try_detect_delimited(["a,b,c", "plain", "text"]) is None


def test_delimited_rejects_inconsistent_columns() -> None:
    # Column count swings too much to be a real table.
    assert _try_detect_delimited(["a,b", "c,d", "e,f,g,h", "i,j,k,l,m"]) is None


def test_delimited_keeps_first_equal_confidence_delimiter() -> None:
    # Comma and semicolon are both consistent; the comma candidate is set first
    # and a later, no-better delimiter does not displace it.
    result = _try_detect_delimited(["a,b;c", "d,e;f", "g,h;i"])
    assert result is not None
    assert result.metadata["delimiter"] == ","


def test_looks_like_prose_distinguishes_sentences_from_rows() -> None:
    # Wordy cells (avg > 3 words/cell) read as prose even without end punctuation.
    assert _looks_like_prose(["the quick brown fox runs, over the lazy dog now"], ",")
    # Short field tuples are real CSV rows, not prose.
    assert not _looks_like_prose(["a,b,c", "1,2,3", "x,y,z"], ",")


# Parsers --------------------------------------------------------------------


def test_parse_csv_and_records() -> None:
    headers, rows = parse_csv(CSV)
    assert headers == ["name", "age", "city"]
    assert rows[0] == ["Alice", "30", "NYC"]
    records = to_records(headers, rows)
    assert records[1] == {"name": "Bob", "age": "25", "city": "LA"}


def test_parse_markdown_table_drops_separator() -> None:
    headers, rows = parse_markdown_table(MARKDOWN)
    assert headers == ["name", "age"]
    assert ["Alice", "30"] in rows
    assert all("---" not in cell for row in rows for cell in row)


def test_parse_tabular_rejects_ragged_fixed_width(monkeypatch) -> None:
    # Rows with differing cell counts can't be zipped under the headers
    # without misattributing columns (#1652) — must pass through.
    import headroom.transforms.tabular_ingest as ti

    monkeypatch.setattr(
        ti,
        "detect_content_type",
        lambda _c: DetectionResult(ContentType.TABULAR, 0.9, {"format": "fixed_width"}),
    )
    ragged = (
        "tool  installed  latest  status\n"
        "rtk  0.42.4  0.43.0  update available\n"
        "rtk  ✓  0.42.4  0.42.4  -  up-to-date"
    )
    assert ti.parse_tabular(ragged) is None


def test_parse_tabular_rejects_ragged_markdown(monkeypatch) -> None:
    import headroom.transforms.tabular_ingest as ti

    monkeypatch.setattr(
        ti,
        "detect_content_type",
        lambda _c: DetectionResult(ContentType.TABULAR, 0.9, {"format": "markdown"}),
    )
    ragged = "| a | b | c |\n| --- | --- | --- |\n| 1 | 2 | 3 |\n| 4 | 5 |"
    assert ti.parse_tabular(ragged) is None


def test_compress_passes_through_ragged_table(monkeypatch) -> None:
    import headroom.transforms.tabular_ingest as ti

    monkeypatch.setattr(
        ti,
        "detect_content_type",
        lambda _c: DetectionResult(ContentType.TABULAR, 0.9, {"format": "fixed_width"}),
    )
    ragged = (
        "tool  installed  latest  status\n"
        "rtk  0.42.4  0.43.0  update available\n"
        "rtk  ✓  0.42.4  0.42.4  -  up-to-date"
    )
    result = TabularCompressor().compress(ragged)
    assert not result.was_modified
    assert result.compressed == ragged


def test_parse_tabular_returns_none_for_non_tabular() -> None:
    assert parse_tabular("just a normal paragraph here") is None


def test_parse_fixed_width() -> None:
    headers, rows = parse_fixed_width("name    age   city\nAlice   30    NYC\nBob     25    LA")
    assert headers == ["name", "age", "city"]
    assert rows[0] == ["Alice", "30", "NYC"]


def test_to_records_empty_headers_returns_empty() -> None:
    assert to_records([], [["a", "b"]]) == []


def test_parse_csv_blank_returns_empty() -> None:
    assert parse_csv("   \n  \n") == ([], [])


def test_parse_markdown_table_too_short_returns_empty() -> None:
    assert parse_markdown_table("| only one row |") == ([], [])


def test_parse_fixed_width_too_short_returns_empty() -> None:
    assert parse_fixed_width("a single line") == ([], [])


def test_parse_tabular_dispatches_fixed_width(monkeypatch) -> None:
    # The detector currently emits only csv/markdown, so drive the fixed_width
    # dispatch branch directly with a stubbed detection result.
    import headroom.transforms.tabular_ingest as ti

    monkeypatch.setattr(
        ti,
        "detect_content_type",
        lambda _c: DetectionResult(ContentType.TABULAR, 0.9, {"format": "fixed_width"}),
    )
    headers, rows, fmt = ti.parse_tabular("name    age\nAlice   30\nBob     25")
    assert fmt == "fixed_width"
    assert headers == ["name", "age"]
    assert rows[0] == ["Alice", "30"]


def test_parse_tabular_none_when_no_data_rows_survive() -> None:
    # Detected as a markdown table, but it is header + separator rows only:
    # nothing survives as a data row, so parse_tabular bails to None.
    assert parse_tabular("| a | b |\n| --- | --- |\n| --- | --- |") is None


def test_compression_ratio_zero_for_empty_original() -> None:
    result = TabularCompressionResult(
        compressed="", original="", was_modified=False, fmt="csv", rows=0, columns=0
    )
    assert result.compression_ratio == 0.0


# Bridge compressor ----------------------------------------------------------


def test_verbose_markdown_compresses() -> None:
    result = TabularCompressor().compress(_verbose_markdown())
    assert result.was_modified
    assert len(result.compressed) < len(result.original)
    assert result.compression_ratio < 1.0
    assert result.fmt == "markdown"


def test_compact_unique_csv_passes_through() -> None:
    # All-unique compact rows have nothing losslessly removable.
    result = TabularCompressor().compress(CSV)
    assert not result.was_modified
    assert result.compressed == CSV


def test_non_tabular_passes_through_unmodified() -> None:
    # Unparseable prose returns the original content untouched.
    text = "just a normal paragraph here"
    result = TabularCompressor().compress(text)
    assert not result.was_modified
    assert result.compressed == text


# Router wiring --------------------------------------------------------------


def test_router_routes_tabular() -> None:
    result = ContentRouter().compress(_verbose_markdown())
    assert result.strategy_used is CompressionStrategy.TABULAR
    assert result.total_compressed_tokens <= result.total_original_tokens


def test_router_caches_tabular_compressor() -> None:
    router = ContentRouter()
    first = router._get_tabular_compressor()
    assert first is router._get_tabular_compressor()  # second call returns the cached instance


def test_router_tabular_passthrough_when_compressor_unavailable(monkeypatch) -> None:
    # Defensive guard: if the tabular compressor can't be constructed, routing to
    # TABULAR leaves content untouched instead of crashing.
    md = _verbose_markdown()
    router = ContentRouter()
    monkeypatch.setattr(router, "_get_tabular_compressor", lambda: None)
    result = router.compress(md)
    assert result.compressed == md
    assert result.tokens_saved == 0


def test_router_respects_disable_flag() -> None:
    # Disabling skips the tabular compressor: content passes through unchanged
    # (the selected strategy label may still read TABULAR, like other disabled
    # compressors).
    md = _verbose_markdown()
    cfg = ContentRouterConfig(enable_tabular_compressor=False)
    result = ContentRouter(cfg).compress(md)
    assert result.compressed == md
    assert result.tokens_saved == 0


# Binary spreadsheet ingestion -----------------------------------------------


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl not installed")
def test_load_and_compress_xlsx(tmp_path) -> None:
    import openpyxl

    from headroom import compress_spreadsheet
    from headroom.transforms.spreadsheet_ingest import load_spreadsheet

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["id", "name", "dept", "status"])
    for i in range(40):
        ws.append([i, f"user_{i}", ["eng", "sales", "ops"][i % 3], "active"])
    wb.create_sheet("Empty")  # should be skipped
    path = tmp_path / "sample.xlsx"
    wb.save(path)

    sheets = load_spreadsheet(path)
    assert list(sheets) == ["Data"]
    assert sheets["Data"].splitlines()[0] == "id,name,dept,status"

    result = compress_spreadsheet(str(path))
    assert result.tokens_after <= result.tokens_before


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl not installed")
def test_compress_spreadsheet_empty_workbook_returns_empty(tmp_path) -> None:
    import openpyxl

    from headroom import compress_spreadsheet

    wb = openpyxl.Workbook()  # one empty sheet, no rows
    path = tmp_path / "empty.xlsx"
    wb.save(path)

    result = compress_spreadsheet(str(path))
    assert result.messages == []
    assert result.tokens_saved == 0


def test_load_spreadsheet_rejects_unknown_extension(tmp_path) -> None:
    from headroom.transforms.spreadsheet_ingest import load_spreadsheet

    bad = tmp_path / "data.txt"
    bad.write_text("a,b\n1,2\n")
    with pytest.raises(ValueError, match="Unsupported"):
        load_spreadsheet(bad)


def test_load_spreadsheet_missing_file(tmp_path) -> None:
    from headroom.transforms.spreadsheet_ingest import load_spreadsheet

    with pytest.raises(FileNotFoundError):
        load_spreadsheet(tmp_path / "nope.xlsx")
